import csv
import math
import os
import re
import uuid
from io import BytesIO, StringIO
from typing import Any

from email_validator import EmailNotValidError, validate_email
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.models import CsvUpload, Lead, LeadStatus, LogType, SentEmail
from app.services.common import is_valid_email, log_activity

EMAIL_IN_TEXT = re.compile(
    r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}"
)

COLUMN_MAP = {
    "company": "company",
    "company name": "company",
    "company_name": "company",
    "organisation": "company",
    "organization": "company",
    "business": "company",
    "client": "company",
    "email": "email",
    "e-mail": "email",
    "email address": "email",
    "email_address": "email",
    "mail": "email",
    "first name": "first_name",
    "firstname": "first_name",
    "first_name": "first_name",
    "last name": "last_name",
    "lastname": "last_name",
    "last_name": "last_name",
    "name": "first_name",
    "full name": "first_name",
    "contact": "first_name",
    "contact name": "first_name",
    "website": "website",
    "web": "website",
    "url": "website",
    "phone": "phone",
    "phone number": "phone",
    "phonenumber": "phone",
    "phone_number": "phone",
    "mobile": "phone",
    "mobile number": "phone",
    "telephone": "phone",
    "tel": "phone",
    "country": "country",
    "industry": "industry",
    "category": "industry",
    "notes / source": "source_url",
    "notes/source": "source_url",
    "notes": "address",
    "source": "source_url",
    "remarks": "address",
    "remark": "address",
}


def _clean_header(name: Any) -> str:
    text = str(name or "").replace("\ufeff", "").strip().lower()
    text = re.sub(r"\s+", " ", text)
    return COLUMN_MAP.get(text, text)


def _decode_csv_text(file_content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return file_content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return file_content.decode("utf-8", errors="replace")


def _read_csv_rows(file_content: bytes) -> list[dict[str, Any]]:
    text = _decode_csv_text(file_content)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(StringIO(text), dialect)
    try:
        raw_headers = next(reader)
    except StopIteration:
        return []

    headers = [_clean_header(h) for h in raw_headers]
    rows: list[dict[str, Any]] = []
    for values in reader:
        if not any(str(v or "").strip() for v in values):
            continue
        row: dict[str, Any] = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            row[header] = values[i] if i < len(values) else None
        rows.append(row)
    return rows


def _read_excel_rows(file_content: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
    ws = wb.active
    row_iter = ws.iter_rows(values_only=True)
    try:
        raw_headers = next(row_iter)
    except StopIteration:
        return []

    headers = [_clean_header(h) for h in raw_headers]
    rows: list[dict[str, Any]] = []
    for values in row_iter:
        if not any(str(v or "").strip() for v in values):
            continue
        row: dict[str, Any] = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            row[header] = values[i] if i < len(values) else None
        rows.append(row)
    return rows


def _read_tabular(file_content: bytes, original_filename: str) -> list[dict[str, Any]]:
    lower = original_filename.lower()
    if lower.endswith((".xlsx", ".xls")):
        if lower.endswith(".xls"):
            raise ValueError("Legacy .xls files are not supported — save as .xlsx")
        return _read_excel_rows(file_content)
    return _read_csv_rows(file_content)


def _validate_email_address(email: str) -> bool:
    try:
        validate_email(email, check_deliverability=False)
        return is_valid_email(email)
    except EmailNotValidError:
        return False


def _extract_email_from_row(row: dict[str, Any]) -> str | None:
    direct = _safe_str(row.get("email"))
    if direct and _validate_email_address(direct.lower()):
        return direct.lower()

    for value in row.values():
        text = _safe_str(value)
        if not text:
            continue
        match = EMAIL_IN_TEXT.search(text)
        if match:
            candidate = match.group(0).lower()
            if _validate_email_address(candidate):
                return candidate
    return None


async def _email_already_sent(db: AsyncSession, email: str) -> bool:
    result = await db.execute(select(SentEmail).where(SentEmail.email == email.lower()))
    return result.scalar_one_or_none() is not None


async def _lead_exists(db: AsyncSession, email: str) -> Lead | None:
    result = await db.execute(select(Lead).where(Lead.email == email.lower()))
    return result.scalar_one_or_none()


async def _lead_exists_by_phone(db: AsyncSession, phone: str) -> Lead | None:
    result = await db.execute(select(Lead).where(Lead.phone == phone))
    return result.scalar_one_or_none()


async def process_csv_upload(
    db: AsyncSession,
    file_content: bytes,
    original_filename: str,
) -> CsvUpload:
    """
    Flexible import: maps common Excel headers (phone number, country, category,
    notes / source, remarks, …). Email can be its own column or found inside notes.
    Company is optional. Phone-only rows are kept as no_email leads.
    """
    os.makedirs(settings.upload_dir, exist_ok=True)
    stored_name = f"{uuid.uuid4().hex}_{original_filename}"
    filepath = os.path.join(settings.upload_dir, stored_name)

    with open(filepath, "wb") as f:
        f.write(file_content)

    rows = _read_tabular(file_content, original_filename)

    if not rows:
        raise ValueError("File is empty — no rows to import")

    upload = CsvUpload(
        filename=stored_name,
        original_filename=original_filename,
        total_rows=len(rows),
    )
    db.add(upload)
    await db.flush()

    valid = duplicate = invalid = 0

    for row in rows:
        email_raw = _extract_email_from_row(row)
        phone = _safe_str(row.get("phone"))
        company = _safe_str(row.get("company"))
        industry = _safe_str(row.get("industry"))
        country = _safe_str(row.get("country"))
        first_name = _safe_str(row.get("first_name"))
        last_name = _safe_str(row.get("last_name"))
        website = _safe_str(row.get("website"))
        source_url = _safe_str(row.get("source_url"))
        address = _safe_str(row.get("address"))

        if not company and industry:
            company = industry

        if not email_raw and not phone and not company and not first_name:
            invalid += 1
            continue

        if email_raw:
            if await _email_already_sent(db, email_raw):
                duplicate += 1
                continue
            existing = await _lead_exists(db, email_raw)
            if existing:
                duplicate += 1
                continue
            status = LeadStatus.VALID
        else:
            if phone:
                existing_phone = await _lead_exists_by_phone(db, phone)
                if existing_phone:
                    duplicate += 1
                    continue
            status = LeadStatus.NO_EMAIL

        lead = Lead(
            company=company,
            email=email_raw,
            first_name=first_name,
            last_name=last_name,
            website=website,
            phone=phone,
            country=country,
            industry=industry,
            source_url=source_url,
            address=address,
            status=status,
        )
        db.add(lead)
        valid += 1

    upload.valid_rows = valid
    upload.duplicate_rows = duplicate
    upload.invalid_rows = invalid

    await log_activity(
        db,
        LogType.CSV_UPLOADED,
        f"CSV/Excel uploaded: {original_filename}",
        f"Imported: {valid}, Duplicates: {duplicate}, Skipped empty: {invalid}",
    )

    return upload


def _safe_str(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    s = str(value).strip()
    return s if s and s.lower() != "nan" else None


def _workbook_to_bytes(wb: Workbook) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


async def export_leads_to_excel(db: AsyncSession, lead_ids: list[int] | None = None) -> bytes:
    query = select(Lead)
    if lead_ids:
        query = query.where(Lead.id.in_(lead_ids))
    result = await db.execute(query)
    leads = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Prospects"
    ws.append(
        [
            "Company",
            "Email",
            "First Name",
            "Last Name",
            "Website",
            "Phone",
            "Country",
            "Status",
            "Created",
        ]
    )
    for lead in leads:
        ws.append(
            [
                lead.company,
                lead.email,
                lead.first_name,
                lead.last_name,
                lead.website,
                lead.phone,
                lead.country,
                lead.status.value if lead.status else None,
                lead.created_at,
            ]
        )
    return _workbook_to_bytes(wb)


async def export_sent_leads(db: AsyncSession, campaign_id: int | None = None) -> bytes:
    query = select(SentEmail)
    if campaign_id:
        query = query.where(SentEmail.campaign_id == campaign_id)
    result = await db.execute(query)
    sent = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Sent Successfully"
    ws.append(["Company", "Email", "Sent Date", "Campaign ID", "Mailbox ID"])
    for row in sent:
        ws.append([row.company, row.email, row.sent_at, row.campaign_id, row.mailbox_id])
    return _workbook_to_bytes(wb)
